# import csv
# import sys
# import time
# from datetime import datetime

# import serial
# from openpyxl import Workbook, load_workbook

# # ---- Configuration ----------------------------------------------------
# SERIAL_PORT = "COM4"        # Windows: "COM3", "COM4", ...
# BAUD_RATE = 9600            # must match Serial.begin(...) in your sketch
# EXCEL_FILE = "arduino_data.xlsx"
# SHEET_NAME = "Attenuator Data Reading"
# COLUMN_HEADERS = ["Time", "Value"]   # add more headers if you send
#                                            # multiple comma-separated values
# # -------------------------------------------------------------------------


# def get_or_create_workbook(path, sheet_name, headers):
#     try:
#         wb = load_workbook(path)
#         ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
#     except FileNotFoundError:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = sheet_name
#         ws.append(headers)
#     return wb, ws


# def main():
#     try:
#         ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)
#     except serial.SerialException as e:
#         print(f"Could not open {SERIAL_PORT}: {e}")
#         print("Check the port name and that no other program (like the Arduino IDE "
#               "Serial Monitor) has it open.")
#         sys.exit(1)

#     # Give the Arduino a moment to reset after the serial connection opens
#     time.sleep(2)

#     wb, ws = get_or_create_workbook(EXCEL_FILE, SHEET_NAME, COLUMN_HEADERS)

#     print(f"Listening on {SERIAL_PORT} at {BAUD_RATE} baud. Logging to {EXCEL_FILE}.")
#     print("Press Ctrl+C to stop.\n")

#     try:
#         while True:
#             line = ser.readline().decode("utf-8", errors="ignore").strip()
#             if not line:
#                 continue

#             timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#             # If the Arduino sends comma-separated values, split them into
#             # separate columns; otherwise log the single value as-is.
#             values = next(csv.reader([line]))
#             row = [timestamp] + values

#             ws.append(row)
#             wb.save(EXCEL_FILE)  # save after every reading so nothing is lost

#             print(row)

#     except KeyboardInterrupt:
#         print("\nStopping... saving final file.")
#     finally:
#         wb.save(EXCEL_FILE)
#         ser.close()
#         print(f"Done. Data saved to {EXCEL_FILE}")


# if __name__ == "__main__":
#     main()

import csv
import sys
import time
from datetime import datetime

import serial
from openpyxl import Workbook, load_workbook

# ---- Configuration ----------------------------------------------------
SERIAL_PORT = "COM4"        # Windows: "COM3", "COM4", ...
                             # Mac/Linux: "/dev/ttyUSB0" or "/dev/ttyACM0"
BAUD_RATE = 9600            # must match Serial.begin(...) in your sketch
EXCEL_FILE = "arduino_data.xlsx"
SHEET_NAME = "Attenuator Data Reading 2"
COLUMN_HEADERS = ["Time", "Value"]   # add more headers if you send
                                           # multiple comma-separated values
# -------------------------------------------------------------------------


def get_or_create_workbook(path, sheet_name, headers):
    try:
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
    return wb, ws


def main():
    try:
        ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)
    except serial.SerialException as e:
        print(f"Could not open {SERIAL_PORT}: {e}")
        print("Check the port name and that no other program (like the Arduino IDE "
              "Serial Monitor) has it open.")
        sys.exit(1)

    # Give the Arduino a moment to reset after the serial connection opens
    time.sleep(2)

    wb, ws = get_or_create_workbook(EXCEL_FILE, SHEET_NAME, COLUMN_HEADERS)

    print(f"Listening on {SERIAL_PORT} at {BAUD_RATE} baud. Logging to {EXCEL_FILE}.")
    print("Press Ctrl+C to stop.\n")

    try:
        while True:
            line = ser.readline().decode("utf-8", errors="ignore").strip()
            if not line:
                continue

            # %f gives microseconds (6 digits); slice to 3 for milliseconds
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]

            # If the Arduino sends comma-separated values, split them into
            # separate columns; otherwise log the single value as-is.
            values = next(csv.reader([line]))
            row = [timestamp] + values

            ws.append(row)
            wb.save(EXCEL_FILE)  # save after every reading so nothing is lost

            print(row)

    except KeyboardInterrupt:
        print("\nStopping... saving final file.")
    finally:
        wb.save(EXCEL_FILE)
        ser.close()
        print(f"Done. Data saved to {EXCEL_FILE}")


if __name__ == "__main__":
    main()